#coding=utf-8
import openpyxl
from openpyxl import load_workbook,Workbook
wb=load_workbook(r'E:\data\table\execl\5m方格叠加\building09.xlsx')
wb1=load_workbook(r'E:\data\GIS数据\网络优化\Heights\changshashi5m.xlsx')
sheet=wb.active
sheet1=wb1.active
a=[]
b=[]
for i in range(1,33382):    
    for j in range(1,61402):
        if sheet['M%d'%i].value==sheet1['A%d'%j].value:
            a.append(j)
        if sheet['N%d'%i].value==sheet1['B%d'%j].value:
            b.append(j)
        c=[val for val in a if val in b]
        if c!=[]:
            sheet['O%d'%i].value=sheet1['C%d'%c[0]].value
            break
           
wb.save(r'E:\data\table\execl\5m方格叠加\buildin09.xlsx')

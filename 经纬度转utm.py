#coding=utf-8

import math
import sys
import openpyxl
from openpyxl import load_workbook,Workbook

# 常量定义
a = 6378137
b = 6356752.3142451
f = (a - b)/a
lonOrigin = 111

sin = math.sin
cos = math.cos
tan = math.tan
sqrt = math.sqrt
radians = math.radians
degrees = math.degrees
pi = math.pi

def LL2UTM_USGS(lat,lon):
    '''
    ** Input：(a, f, lat, lon, lonOrigin, FN)
    ** a 椭球体长半轴
    ** f 椭球体扁率 f=(a-b)/a 其中b代表椭球体的短半轴
    ** lat 经过UTM投影之前的纬度（角度）
    ** lon 经过UTM投影之前的经度（角度）
    ** lonOrigin 中央经度线（角度）
    ** FN 纬度起始点，北半球为0，南半球为10000000.0m
    ---------------------------------------------
    ** Output:(UTMNorthing, UTMEasting)
    ** UTMNorthing 经过UTM投影后的纬度方向的坐标
    ** UTMEasting 经过UTM投影后的经度方向的坐标
    ---------------------------------------------
    ** 功能描述：经纬度坐标投影为UTM坐标，采用美国地理测量部(USGS)提供的公式
    ** 作者： Ace Strong
    ** 单位： CCA NUAA
    ** 创建日期：2008年7月19日
    ** 版本：1.0
    ** 本程序实现的公式请参考
    ** "Coordinate Conversions and Transformations including Formulas" p35.
    ** & http://www.uwgb.edu/dutchs/UsefulData/UTMFormulas.htm
    '''

    # e表示WGS84第一偏心率,eSquare表示e的平方    
    eSquare = 2*f - f*f
    k0 = 0.9996
# 确保longtitude位于-180.00----179.9之间
    lonTemp = (lon+180)-int((lon+180)/360)*360-180
    latRad = radians(lat)
    lonRad = radians(lonTemp)
    lonOriginRad = radians(lonOrigin)
    e2Square = (eSquare)/(1-eSquare)

    V = a/sqrt(1-eSquare*sin(latRad)**2)
    T = tan(latRad)**2
    C = e2Square*cos(latRad)**2
    A = cos(latRad)*(lonRad-lonOriginRad)
    M = a*((1-eSquare/4-3*eSquare**2/64-5*eSquare**3/256)*latRad
    -(3*eSquare/8+3*eSquare**2/32+45*eSquare**3/1024)*sin(2*latRad)
    +(15*eSquare**2/256+45*eSquare**3/1024)*sin(4*latRad)
    -(35*eSquare**3/3072)*sin(6*latRad))

    # x
    UTMEasting = k0*V*(A+(1-T+C)*A**3/6
    + (5-18*T+T**2+72*C-58*e2Square)*A**5/120)+ 500000.0
    # y
    UTMNorthing = k0*(M+V*tan(latRad)*(A**2/2+(5-T+9*C+4*C**2)*A**4/24
    +(61-58*T+T**2+600*C-330*e2Square)*A**6/720))
    # 南半球纬度起点为10000000.0m
    UTMNorthing += 0
    return (UTMEasting, UTMNorthing)

wb=load_workbook(r'E:\data\table\execl\5m方格\coverfacemax.xlsx')
sheetnames = wb.get_sheet_names() #获得表名称
ws = wb.get_sheet_by_name(sheetnames[0]) #获取第一张表
wb1=load_workbook(r"E:\data\table\execl\5m方格叠加\coverfacemax.xlsx")
sheet=wb1.active
for i in range(1,150000):
    if not(ws.cell(row = i,column = 4).value):
        break
    else:
        x=ws.cell(row = i,column = 3).value/10000000
        y=ws.cell(row = i,column = 2).value/10000000   
        zuobiao=LL2UTM_USGS(x,y)    
        sheet['K%d'%i].value=round(zuobiao[0],1)
        sheet['L%d'%i].value=round(zuobiao[1],1)    
wb1.save(r"E:\data\table\execl\5m方格叠加\coverfacemax.xlsx")




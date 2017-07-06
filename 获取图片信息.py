#!/usr/bin/python
# -*- coding:utf-8 -*-
from PIL import Image
import os
import os.path
from openpyxl import Workbook
dirs = input('输入文件目录：')

def exif_data(filename):
	img = Image.open(filename)
	exif = img._getexif()
	times = exif[36867]
	xyz = exif[34853]
	date = times[:10]
	dates = date.replace(':','/')
	time = times[10:]
	datetime  = dates+time
	longitude = xyz[4]
	latitude = xyz[2]
	Altitude = xyz[6]
	longitudes = longitude[0][0]+longitude[1][0]/60+longitude[2][0]/(longitude[2][1]*3600)
	latitudes = latitude[0][0]+latitude[1][0]/60+latitude[2][0]/(latitude[2][1]*3600)
	Altitudes = Altitude[0]/Altitude[1]
	return datetime,longitudes,latitudes,Altitudes
wb = Workbook()
sheet = wb.active
for root,dir,files in os.walk(dirs):
	for j, file in enumerate(files):
		if file.endswith('.jpg'):
			j= j + 1
			roots = os.path.join(root,file)
			datetime, longitudes, latitudes, Altitudes = exif_data(roots)
			sheet.cell(row=j, column=1).value = file
			sheet.cell(row=j, column=2).value= datetime
			sheet.cell(row=j, column=3).value= longitudes
			sheet.cell(row=j, column=4).value= latitudes
			sheet.cell(row=j, column=5).value= Altitudes
			# print(file, times, longitudes, latitudes, Altitudes)
wb.save(os.path.join(dirs,'image.xlsx'))


